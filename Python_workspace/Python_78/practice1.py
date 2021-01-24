import openpyxl,requests
# 读取数据
def read_data(filename,sheetname):
    wb = openpyxl.load_workbook(filename)
    sheet = wb[sheetname]
    max_row = sheet.max_row
    cases = []
    for i in range(2,max_row+1):
        dict1 = dict(
        case_id = sheet.cell(row=i,column = 1).value,
        url = sheet.cell(row=i,column = 5).value,
        data = sheet.cell(row=i,column = 6).value,
        expected = sheet.cell(row=i,column = 7).value)
        cases.append(dict1)
    return cases


# 执行接口请求
def api_func(url_api,data_api):
    header = {'X-Lemonban-Media-Type':'lemonban.v2'}
    response=requests.post(url=url_api,json=data_api,headers=header)
    result = response.json()
    return result

# 写入数据
def write_data(filename,sheetname,row,column,final_result):
    wb = openpyxl.load_workbook(filename)
    sheet = wb[sheetname]
    sheet.cell(row=row,column=column).value= final_result
    wb.save(filename)

# 执行测试
'''
{'case_id': 1, 'url': 'http://api.lemonban.com/futureloan/member/register', 'data': '{"mobile_phone":"13552440101","pwd":"12345678","type":1,"reg_name":"34254sdfs"}', 'expected': '{"code": 0, "msg": "OK"}'}
'''
def execute_func(filename,sheetname):
    cases = read_data(filename,sheetname)
    for case in cases:
        case_id = case.get('case_id')
        url = case.get('url')
        data = case.get('data')
        data = eval(data)
        expected_result = case.get('expected')
        expected_result = eval(expected_result)
        expected_msg = expected_result.get('msg')
        real_result = api_func(url_api=url, data_api=data)
        real_msg = real_result.get('msg')
        print('预期执行结果为：{}'.format(expected_msg))
        print('实际执行结果为：{}'.format(real_msg))
        if expected_msg == real_msg:
            final_result = 'Passed'
            print('第{}条测试用例通过'.format(case_id))
        else:
            final_result = 'Failed'
            print('第{}条测试用例不通过'.format(case_id))
        print('*'*20)
        write_data(filename,sheetname,case_id+1,8,final_result)

execute_func('test_case_api.xlsx','login')






