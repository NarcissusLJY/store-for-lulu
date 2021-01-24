import openpyxl
'''
读取数据
1.获取工作簿 openpyxl.loadworkbook('')
2.获取表单 wb[]
3.获取单元格

# 写入数据
1.赋值
2.wb.save()
'''
# wb = openpyxl.load_workbook("test_case_api.xlsx")  # 获取工作簿
# sheet = wb['register']   # 获取表单
# cell = sheet.cell(row=1,column=1)  # 获取单元格
# print(cell.value)
# cell.value = '用例编号'  # 对这个单元格进行重新赋值
# print(cell.value)
# wb.save('test_case_api.xlsx')

# 读取测试用例
wb = openpyxl.load_workbook('test_case_api.xlsx')
sheet = wb['register']
max_row = sheet.max_row
cases = []
for i in range(2,max_row+1):
    dict1 = dict(
    case_id = sheet.cell(row=i,column=1,).value,
    header = sheet.cell(row=i,column=4,).value,
    url = sheet.cell(row=i,column=6,).value,
    data = sheet.cell(row=i,column=7,).value,
    expect = sheet.cell(row=i,column=8,).value)
    cases.append(dict1)
print(cases)



