import openpyxl  #或者 from openpyxl import loadworkbook
import pprint
import requests
'''
读取数据
1.获取表格
2.获取表单sheet
3.获取单元格
数据写入
1.赋值
2.wb.save

wb = openpyxl.load_workbook('test_case_api.xlsx')
sheet = wb['register']
cell = sheet.cell(row=1,column=1)
print(cell.value)
cell.value = "用例编号" 
wb.save('test_case_api.xlsx')

'''



# 获取用例
def read_data(filename,sheetname):
    wb = openpyxl.load_workbook(filename)
    sheet = wb[sheetname]
    # 获取最大行数
    max_row = sheet.max_row
    cases = []
    for i in range(2,max_row+1):
        case = dict(
        case_id = sheet.cell(row=i,column=1).value,
        url = sheet.cell(row=i,column=5).value,
        data = sheet.cell(row=i,column=6).value,
        expect_result = sheet.cell(row=i,column=7).value)
        cases.append(case)
    return cases

# 调用函数不影响其他需要导入函数的文件，在原文件中调用函数前使用main函数
if __name__ == '__main__':

    final_reslut=read_data('test_case_api.xlsx','register')
    print(final_reslut)


# 发送接口请求
def api_func(url_api,data_api):
    header = {'X-Lemonban-Media-Type':'lemonban.v2','Content-Type':'application/json'}
    response=requests.post(url=url_api,json=data_api,headers=header)
    result=response.json()
    return result

if __name__ == '__main__':

    url = 'http://api.lemonban.com/futureloan/member/register'
    data = {"mobile_phone":"13512440101","pwd":"12345678","type":1,"reg_name":"34254sdfs"}
    resp = api_func(url,data)
    pprint.pprint(resp)

# # 执行用例
# def perform_func(filename,sheetname):
#     cases = read_data(filename,sheetname)
#     # print(cases)
#
# #     result = requests.post(url=url,, json=)
# #     response = result.json()
# #     return response
# #
# #
# re = perform_func('test_case_api.xlsx','register')
# # pprint.pprint(re)


# 数据写入的函数
def write_result(filename,sheetname,final_result,row,column):
    wb = openpyxl.load_workbook(filename)
    sheet = wb[sheetname]
    sheet.cell(row=row,column=column).value = final_result
    wb.save(filename)





